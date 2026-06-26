from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class SheetProfile:
    sheet_name: str
    sheet_state: str
    max_row: int
    max_column: int
    has_auto_filter: bool
    freeze_panes: str | None
    merged_cell_count: int


@dataclass
class WorkbookProfile:
    workbook_path: str
    workbook_name: str
    file_size_mb: float
    sheet_count: int
    visible_sheet_count: int
    hidden_sheet_count: int
    sheet_names: list[str]
    data_sheet_exists: bool
    data_sheet_headers: list[str]
    sheets: list[SheetProfile]


def _clean_cell_value(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    text = " ".join(text.split())
    return text


def _detect_header_row(ws, max_scan_rows: int = 20) -> list[str]:
    """
    Detect a likely header row by scanning the first rows and choosing
    the row with the highest number of non-empty cells.
    """
    best_values: list[str] = []
    best_non_empty_count = 0

    max_row_to_scan = min(ws.max_row or 0, max_scan_rows)

    for row in ws.iter_rows(min_row=1, max_row=max_row_to_scan, values_only=True):
        cleaned_values = [_clean_cell_value(value) for value in row]
        non_empty_values = [value for value in cleaned_values if value]

        if len(non_empty_values) > best_non_empty_count:
            best_non_empty_count = len(non_empty_values)
            best_values = non_empty_values

    return best_values


def profile_workbook(workbook_path: str | Path) -> WorkbookProfile:
    """
    Profile an Excel workbook without transforming business data.

    This function does not extract part numbers yet.
    It only inspects workbook structure.
    """
    path = Path(workbook_path)

    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Expected .xlsx or .xlsm workbook, got: {path.suffix}")

    workbook = load_workbook(filename=path, read_only=False, data_only=False)

    sheet_profiles: list[SheetProfile] = []

    for ws in workbook.worksheets:
        merged_cell_count = len(ws.merged_cells.ranges) if hasattr(ws, "merged_cells") else 0

        sheet_profiles.append(
            SheetProfile(
                sheet_name=ws.title,
                sheet_state=ws.sheet_state,
                max_row=ws.max_row or 0,
                max_column=ws.max_column or 0,
                has_auto_filter=bool(ws.auto_filter and ws.auto_filter.ref),
                freeze_panes=str(ws.freeze_panes) if ws.freeze_panes else None,
                merged_cell_count=merged_cell_count,
            )
        )

    sheet_names = workbook.sheetnames
    visible_sheet_count = sum(1 for sheet in sheet_profiles if sheet.sheet_state == "visible")
    hidden_sheet_count = len(sheet_profiles) - visible_sheet_count

    data_sheet_headers: list[str] = []
    data_sheet_exists = "Data" in sheet_names

    if data_sheet_exists:
        data_ws = workbook["Data"]
        data_sheet_headers = _detect_header_row(data_ws)

    return WorkbookProfile(
        workbook_path=str(path),
        workbook_name=path.name,
        file_size_mb=round(path.stat().st_size / (1024 * 1024), 2),
        sheet_count=len(sheet_profiles),
        visible_sheet_count=visible_sheet_count,
        hidden_sheet_count=hidden_sheet_count,
        sheet_names=sheet_names,
        data_sheet_exists=data_sheet_exists,
        data_sheet_headers=data_sheet_headers,
        sheets=sheet_profiles,
    )


def profile_to_dict(profile: WorkbookProfile) -> dict[str, Any]:
    return asdict(profile)
